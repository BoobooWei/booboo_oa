# -*- coding:utf8 -*-
"""
Created on:
@author: BoobooWei
Email: rgweiyaping@hotmail.com
Version: V.19.03.09.0
Description:
Help:
"""
import xlsxwriter
from openpyxl import load_workbook
import sys

reload(sys)
sys.setdefaultencoding('utf8')


class GetMyExcel:
    """
    读取excel数据
    """

    def __init__(self, excel):
        # read excel
        self.wb = load_workbook(excel)
        # get all sheet names
        # [u'Sheet1', u'Sheet2', u'Sheet3'}
        self.sheetnames = self.wb.sheetnames

    def get_sheet_data(self):
        lines = []
        for sheetname in self.sheetnames:
            sheet = self.wb.get_sheet_by_name(sheetname)
            # get row num
            row_num = sheet.max_row
            # get column num
            column_num = sheet.max_column
            # 获取标题 title = ['a','b']
            title = map(lambda x: x.value, sheet['1'])

            # 获取数据
            for row in range(2, row_num + 1):
                lines.append(map(lambda x: x.value, sheet[row]))
        return (title, lines)

    def filter_column(self, title, lines, except_list):
        try:
            _title = filter(lambda x: title.index(x) not in except_list, title)
            _lines = map(lambda line: filter(lambda x: line.index(x) not in except_list, line), lines)
        except Exception:
            _title = []
            _lines = []
        return (_title, _lines)


class CreateMyExcel:
    """
    创建excel表格
    """

    def __init__(self, excel):
        # Create an new Excel file.
        self.workbook = xlsxwriter.Workbook(excel)

    def create_new_sheet(self):
        # add a worksheet.
        worksheet = self.workbook.add_worksheet()
        return worksheet

    def insert_data(self, worksheet, title, lines):
        # insert title
        for i in range(len(title)):
            # 第一1行，每个列分别写入指定的数据
            worksheet.write(0, i, title[i])
        # insert row
        row = 1
        for line in lines:
            for column in range(len(line)):
                worksheet.write(row, column, line[column])
            row = row + 1

    def close_excel(self):
        self.workbook.close()


def starup(**parmas):
    input_file = parmas["input"]
    output_file = parmas["output"]
    # 将三个sheet合并成一个sheet,过滤需要的列，写入新文件
    get_api = GetMyExcel(input_file)
    _title, _lines = get_api.get_sheet_data()

    except_list = range(14)
    title, lines = get_api.filter_column(_title, _lines, except_list)
    api = CreateMyExcel(output_file)
    api.insert_data(api.create_new_sheet(), title, lines)
    api.close_excel()


if __name__ == "__main__":
    items = [
        {"input": "20190313091400007.xlsx",
         "output": "20190313091400007_end.xlsx"
         },
        {"input": "20190313091600016.xlsx",
         "output": "20190313091600016_ing.xlsx"
         },
    ]
    for params in items:
        starup(**params)
